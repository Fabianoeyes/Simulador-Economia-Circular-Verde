import io
import hashlib
import traceback
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter
from xlcalculator import Evaluator, model, xltypes
from xlcalculator.xlfunctions import func_xltypes, logical, xl

# =========================
# STREAMLIT CONFIG + HEARTBEAT
# =========================
st.set_page_config(page_title="Simulador Economia Circular Verde", layout="wide")

st.title("🚀 Simulador Economia Circular Verde")
st.write("✅ Streamlit está renderizando. (Se você vê isso, o servidor está OK)")
st.caption("Se algo falhar, o erro aparecerá aqui na tela com detalhes.")

# =========================
# CONFIG
# =========================
# ⚠️ Ajuste aqui se o nome da aba for diferente (o app vai mostrar as abas caso não encontre)
MAIN_SHEET = "Simulador Eco Circ Verde"

# Arquivos preferidos
PREFERRED_FILES = [
    "simulador.xlsx",
    "Cópia de Simulador Economia Circular Verde (v.27.03.2025) (2).xlsx",
]

# Ajuste depois para os KPIs reais (células)
OUTPUT_CELLS = {
    "💰 Economia Total": f"{MAIN_SHEET}!M12",
    "📈 ROI": f"{MAIN_SHEET}!M13",
    "🌱 Pontos Ecoa": f"{MAIN_SHEET}!M17",
    "🌍 Impacto": f"{MAIN_SHEET}!M18",
}

# =========================
# HELPERS
# =========================
def find_workbook_in_cwd() -> Path | None:
    """Procura um .xlsx válido na pasta atual, priorizando nomes conhecidos."""
    cwd = Path(".").resolve()

    # 1) prioriza nomes conhecidos
    for name in PREFERRED_FILES:
        p = cwd / name
        if p.exists() and p.is_file() and not p.name.startswith("~$"):
            return p

    # 2) fallback: primeiro xlsx válido encontrado
    for p in cwd.glob("*.xlsx"):
        if p.is_file() and not p.name.startswith("~$"):
            return p

    return None


def is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def is_probably_input_cell(cell) -> bool:
    """
    Heurística para detectar células de input:
    - não é fórmula
    - tem valor
    - e tem fill sólido com theme 7 (ajuste se precisar)
    """
    v = cell.value
    if v is None or v == "":
        return False
    if is_formula(v):
        return False

    fill = cell.fill
    if fill and fill.patternType == "solid" and fill.fgColor and fill.fgColor.type == "theme":
        if fill.fgColor.theme == 7:
            return True

    return False


def discover_inputs(wb: openpyxl.Workbook, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Aba '{sheet_name}' não encontrada.\n"
            f"Abas disponíveis: {wb.sheetnames}"
        )

    ws = wb[sheet_name]

    inputs = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            if not is_probably_input_cell(cell):
                continue

            addr = f"{sheet_name}!{get_column_letter(c)}{r}"

            # tenta label na coluna B da mesma linha
            label = ws.cell(r, 2).value
            label = str(label).strip() if label else addr

            inputs.append(
                {
                    "label": label,
                    "address": addr,
                    "default": cell.value,
                    "row": r,
                    "col": c,
                }
            )

    inputs.sort(key=lambda x: (x["row"], x["col"]))
    return inputs


def coerce_value(v):
    """Converte strings numéricas PT-BR (vírgula) para float quando fizer sentido."""
    if v is None:
        return None
    if isinstance(v, (int, float, bool)):
        return v

    if isinstance(v, str):
        s = v.strip()
        if s.lower() in ("true", "false"):
            return s.lower() == "true"

        if "," in s:
            s2 = s.replace(".", "").replace(",", ".")
        else:
            s2 = s

        try:
            if any(ch.isdigit() for ch in s2):
                return float(s2)
        except Exception:
            return v

    return v


def safe_eval(evaluator: Evaluator, addr: str):
    try:
        return _unwrap_excel_value(evaluator.evaluate(addr))
    except Exception as e:
        return f"Erro: {e}"


def _unwrap_excel_value(val):
    """Converte ExcelTypes / Arrays do xlcalculator em valores Python simples."""
    if isinstance(val, func_xltypes.Expr):
        val = val()
    if isinstance(val, func_xltypes.Array):
        flat = list(val.values.flat)
        return _unwrap_excel_value(flat[0]) if flat else None
    if hasattr(val, "value") and not isinstance(val, (int, float, str, bool)):
        try:
            return val.value
        except Exception:
            pass
    return val


# Monkeypatch do IF para lidar com Arrays de 1x1 e ramos não-callable
@xl.register()
@xl.validate_args
def IF_SAFE(
    logical_test: func_xltypes.XlExpr,
    value_if_true: func_xltypes.XlExpr = True,
    value_if_false: func_xltypes.XlExpr = False,
):
    cond = _unwrap_excel_value(logical_test)
    true_fn = value_if_true if callable(value_if_true) else lambda: value_if_true
    false_fn = value_if_false if callable(value_if_false) else lambda: value_if_false
    chosen = true_fn if bool(cond) else false_fn
    return _unwrap_excel_value(chosen())


logical.IF = IF_SAFE
xl.FUNCTIONS["IF"] = IF_SAFE


def build_model_from_workbook(wb: openpyxl.Workbook) -> model.Model:
    """
    Constrói um Model manualmente (mais leve que ModelCompiler para esta planilha).
    - Só cria ranges quando há ":" (intervalos reais), evitando Arrays 1x1.
    """
    mdl = model.Model()

    for ws in wb.worksheets:
        sheet = ws.title
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue

                addr = f"{sheet}!{cell.coordinate}"
                has_formula = isinstance(v, str) and v.startswith("=")
                xl_cell = xltypes.XLCell(addr, value=None if has_formula else v)

                if has_formula:
                    xl_cell.formula = xltypes.XLFormula(
                        v, sheet_name=sheet, reference=addr
                    )
                    for term in xl_cell.formula.terms:
                        if ":" in term and term not in mdl.ranges:
                            mdl.ranges[term] = xltypes.XLRange(term, term)

                mdl.cells[addr] = xl_cell
                if xl_cell.formula:
                    mdl.formulae[addr] = xl_cell.formula

    mdl.build_code()
    return mdl


@st.cache_resource(hash_funcs={bytes: lambda b: hashlib.sha256(b).hexdigest()})
def load_engine_from_bytes(xlsx_bytes: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=False)
    mdl = build_model_from_workbook(wb)
    return Evaluator(mdl)


@st.cache_data
def cached_inputs(xlsx_bytes: bytes, sheet_name: str):
    with io.BytesIO(xlsx_bytes) as bio:
        wb = openpyxl.load_workbook(bio, data_only=False)
    return discover_inputs(wb, sheet_name)


# =========================
# UI: UPLOAD (OPCIONAL) + XLSX PATH
# =========================
with st.expander("📤 (Opcional) Trocar planilha por upload", expanded=False):
    uploaded = st.file_uploader(
        "Envie um .xlsx (não envie o arquivo que começa com ~$)",
        type=["xlsx"],
    )
    st.info("Se você não fizer upload, o app usa o .xlsx encontrado na pasta do projeto.")

xlsx_bytes: bytes | None = None
xlsx_name: str | None = None
xlsx_path: Path | None = None

if uploaded is not None:
    xlsx_bytes = uploaded.getvalue()
    xlsx_name = uploaded.name
else:
    xlsx_path = find_workbook_in_cwd()
    if xlsx_path:
        xlsx_bytes = xlsx_path.read_bytes()
        xlsx_name = xlsx_path.name

if xlsx_bytes is None or xlsx_name is None:
    st.error(
        "❌ Não encontrei nenhum arquivo .xlsx válido na pasta do projeto. "
        "Coloque o Excel junto do app.py ou faça upload."
    )
    st.stop()

if xlsx_name.startswith("~$"):
    st.error(
        "❌ Você selecionou um arquivo temporário do Excel (começa com '~$'). "
        "Feche o Excel e use o arquivo real."
    )
    st.stop()

st.success(f"📄 Planilha selecionada: **{xlsx_name}**")

# =========================
# DEBUG PANEL
# =========================
with st.expander("🛠️ Debug (ver detalhes)", expanded=False):
    st.write("📍 Pasta atual:", str(Path('.').resolve()))
    if xlsx_path:
        st.write("📍 Arquivo XLSX:", str(xlsx_path.resolve()))
    else:
        st.write("📍 Arquivo XLSX (upload):", xlsx_name)

    try:
        wb_dbg = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=False)
        st.write("📄 Abas encontradas:", wb_dbg.sheetnames)
    except Exception:
        st.error("Falha ao abrir o Excel via openpyxl (apenas leitura).")
        st.code(traceback.format_exc())

# =========================
# LOAD ENGINE (xlcalculator) COM TRY/EXCEPT
# =========================
st.write("Debug: antes de carregar engine do Excel (xlcalculator)")

try:
    engine = load_engine_from_bytes(xlsx_bytes)
    st.success("✅ Engine do Excel carregada (xlcalculator).")
except Exception:
    st.error("❌ Falha ao carregar engine do Excel (xlcalculator).")
    st.code(traceback.format_exc())
    st.stop()

# =========================
# DISCOVER INPUTS COM TRY/EXCEPT
# =========================
st.write("Debug: antes de descobrir inputs na planilha")

try:
    inputs = cached_inputs(xlsx_bytes, MAIN_SHEET)
    st.success(f"✅ Inputs descobertos: {len(inputs)}")
except Exception:
    st.error("❌ Falha ao ler a planilha / aba / inputs.")
    st.code(traceback.format_exc())
    st.stop()

if not inputs:
    st.warning(
        "⚠️ Não encontrei inputs automaticamente (pela cor/estilo). "
        "Sua planilha pode não usar theme 7 para inputs. "
        "Me diga qual aba/cor ou mande print da área de inputs que eu ajusto."
    )
    st.stop()

# =========================
# RENDER INPUTS
# =========================
df = pd.DataFrame(inputs)
df["value"] = df["default"]

st.subheader("✍️ Entradas (editáveis)")
st.caption("Edite os campos. Clique em **Calcular** para atualizar os KPIs.")

edited = st.data_editor(
    df[["label", "address", "value"]],
    use_container_width=True,
    num_rows="fixed",
    column_config={
        "label": st.column_config.TextColumn("Campo"),
        "address": st.column_config.TextColumn("Célula (Excel)"),
        "value": st.column_config.TextColumn("Valor"),
    },
)

col1, col2, col3 = st.columns([1, 1, 3])
with col1:
    calc = st.button("🧮 Calcular", type="primary")
with col2:
    reset = st.button("↩️ Resetar")
with col3:
    st.info("Dica: números podem ser digitados como 1000 ou 1.000,00 (pt-br).")

if reset:
    st.rerun()

# =========================
# CALC + OUTPUTS
# =========================
if calc:
    st.write("Debug: aplicando inputs...")

    edited2 = edited.copy()
    edited2["value"] = edited2["value"].apply(coerce_value)

    for _, row in edited2.iterrows():
        engine.set_cell_value(row["address"], row["value"])

    st.subheader("📌 KPIs")
    cols = st.columns(4)
    i = 0
    for name, addr in OUTPUT_CELLS.items():
        val = safe_eval(engine, addr)
        with cols[i % 4]:
            st.metric(name, str(val))
        i += 1

    st.divider()
    st.subheader("🧾 Logs / Erros")
    st.caption("Se algum KPI der erro, normalmente é fórmula não suportada pelo motor.")
    st.write("Se aparecer 'Erro:', me diga o KPI/célula e eu ajusto a estratégia.")
